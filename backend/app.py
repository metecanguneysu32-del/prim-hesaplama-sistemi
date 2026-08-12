from flask import Flask, jsonify, request, render_template, redirect, url_for
from pathlib import Path
import sys
import csv
import io

from openpyxl import load_workbook


# ============================================================
# DOSYA YOLLARI
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
FRONTEND_DIR = PROJECT_DIR / "frontend"


# ============================================================
# DATABASE
# ============================================================

sys.path.insert(0, str(BASE_DIR))

from database import (
    init_database,
    get_connection,
    search_stores,
    search_personnel,
    create_store,
    create_personnel,
    create_sale,
    create_corporate_sale,
    create_instore_sale,
    save_target
)


# ============================================================
# FLASK
# ============================================================

app = Flask(
    __name__,
    template_folder=str(FRONTEND_DIR),
    static_folder=str(FRONTEND_DIR),
    static_url_path="/static"
)

init_database()


# ============================================================
# YARDIMCI FONKSİYONLAR
# ============================================================

def clean_value(value):
    if value is None:
        return ""

    return str(value).strip()


def normalize_number(value):
    """
    Excel'den gelen satış/hedef tutarlarını sayıya çevirir.

    Desteklenen örnekler:

    125000
    125.000
    125.000,50
    125000,50
    125,000.50
    125000.50
    125.000 TL
    125.000,50 ₺
    """

    if value is None:
        return 0.0

    # Excel hücresi zaten sayı ise
    if isinstance(value, (int, float)):
        return float(value)

    value = clean_value(value)

    if not value:
        return 0.0

    value = value.replace("₺", "")
    value = value.replace("TL", "")
    value = value.replace("tl", "")
    value = value.replace("Tl", "")
    value = value.replace(" ", "")

    # Hem nokta hem virgül varsa
    if "," in value and "." in value:

        # Türkçe format:
        # 125.000,50
        if value.rfind(",") > value.rfind("."):
            value = value.replace(".", "")
            value = value.replace(",", ".")

        # İngilizce format:
        # 125,000.50
        else:
            value = value.replace(",", "")

    # Sadece virgül varsa
    elif "," in value:
        value = value.replace(",", ".")

    # Birden fazla nokta varsa
    elif value.count(".") > 1:
        value = value.replace(".", "")

    try:
        return float(value)

    except ValueError:
        raise ValueError(
            f"Geçersiz sayısal değer: {value}"
        )


# ============================================================
# EXCEL / CSV DOSYA OKUMA
# ============================================================

def read_uploaded_file(file):
    """
    CSV ve XLSX dosyalarını okur.

    XLSX:
        openpyxl kullanılır.

    CSV:
        UTF-8 veya CP1254 okunur.

    XLS:
        Eski Excel formatıdır.
        xlrd kurulu değilse anlaşılır hata verir.
    """

    if file is None:
        raise ValueError(
            "Dosya gönderilmedi."
        )

    filename = clean_value(
        file.filename
    )

    if not filename:
        raise ValueError(
            "Dosya adı bulunamadı."
        )

    extension = Path(
        filename
    ).suffix.lower()

    # --------------------------------------------------------
    # XLSX
    # --------------------------------------------------------

    if extension == ".xlsx":

        try:

            file.stream.seek(0)

            workbook = load_workbook(
                filename=file.stream,
                read_only=True,
                data_only=True
            )

            worksheet = workbook.active

            rows = worksheet.iter_rows(
                values_only=True
            )

            try:
                headers = next(rows)

            except StopIteration:
                workbook.close()

                raise ValueError(
                    "Excel dosyası boş."
                )

            headers = [
                clean_value(header)
                for header in headers
            ]

            result = []

            for row in rows:

                # Tamamen boş satırı atla
                if all(
                    value is None
                    or clean_value(value) == ""
                    for value in row
                ):
                    continue

                item = {}

                for index, header in enumerate(headers):

                    if not header:
                        continue

                    if index < len(row):
                        item[header] = row[index]

                    else:
                        item[header] = None

                result.append(item)

            workbook.close()

            if not result:
                raise ValueError(
                    "Excel dosyasında veri satırı bulunamadı."
                )

            return result

        except Exception as error:

            raise ValueError(
                f"XLSX dosyası okunamadı: {error}"
            )


    # --------------------------------------------------------
    # XLS
    # --------------------------------------------------------

    if extension == ".xls":

        raise ValueError(
            "XLS dosyası desteklenmesi için "
            "xlrd paketi gereklidir. "
            "Şimdilik Excel dosyanızı XLSX olarak kaydedip "
            "tekrar yükleyin."
        )


    # --------------------------------------------------------
    # CSV
    # --------------------------------------------------------

    if extension == ".csv":

        raw = file.read()

        if not raw:
            raise ValueError(
                "Dosya boş."
            )

        try:
            text = raw.decode(
                "utf-8-sig"
            )

        except UnicodeDecodeError:

            text = raw.decode(
                "cp1254"
            )

        return list(
            csv.DictReader(
                io.StringIO(text)
            )
        )


    # --------------------------------------------------------
    # DESTEKLENMEYEN FORMAT
    # --------------------------------------------------------

    raise ValueError(
        "Desteklenmeyen dosya formatı. "
        "Lütfen XLSX veya CSV dosyası yükleyin."
    )


# ============================================================
# SÜTUN DEĞERİ BUL
# ============================================================

def get_row_value(row, *column_names):
    """
    Excel sütun isimlerini farklı yazımlarıyla bulur.
    """

    normalized_row = {}

    for key, value in row.items():

        if key is None:
            continue

        normalized_key = (
            str(key)
            .strip()
            .lower()
        )

        normalized_key = (
            normalized_key
            .replace("ı", "i")
            .replace("İ", "i")
            .replace("ş", "s")
            .replace("Ş", "s")
            .replace("ğ", "g")
            .replace("Ğ", "g")
            .replace("ü", "u")
            .replace("Ü", "u")
            .replace("ö", "o")
            .replace("Ö", "o")
            .replace("ç", "c")
            .replace("Ç", "c")
        )

        normalized_row[
            normalized_key
        ] = value

    for column_name in column_names:

        normalized_name = (
            str(column_name)
            .strip()
            .lower()
        )

        normalized_name = (
            normalized_name
            .replace("ı", "i")
            .replace("İ", "i")
            .replace("ş", "s")
            .replace("Ş", "s")
            .replace("ğ", "g")
            .replace("Ğ", "g")
            .replace("ü", "u")
            .replace("Ü", "u")
            .replace("ö", "o")
            .replace("Ö", "o")
            .replace("ç", "c")
            .replace("Ç", "c")
        )

        if normalized_name in normalized_row:
            return normalized_row[
                normalized_name
            ]

    return None


# ============================================================
# ANA SAYFA
# ============================================================

@app.route("/")
def index():

    return redirect(
        url_for("dashboard")
    )


# ============================================================
# DASHBOARD
# ============================================================

@app.route("/dashboard")
def dashboard():

    dashboard_file = (
        FRONTEND_DIR /
        "dashboard.html"
    )

    if dashboard_file.exists():

        return render_template(
            "dashboard.html"
        )

    return """
    <!DOCTYPE html>
    <html lang="tr">

    <head>

        <meta charset="UTF-8">

        <title>
            Prim Hesaplama Sistemi
        </title>

    </head>

    <body>

        <h1>
            Prim Hesaplama Sistemi
        </h1>

        <p>
            Dashboard dosyası bulunamadı.
        </p>

    </body>

    </html>
    """


# ============================================================
# VERİ AKTARMA
# ============================================================

@app.route("/import")
@app.route("/veri-aktarma")
def import_page():

    return render_template(
        "import.html"
    )


# ============================================================
# KURUMSAL SATIŞ SAYFASI
# ============================================================

@app.route("/corporate-sales")
def corporate_sales_page():

    return render_template(
        "corporate_sales.html"
    )


# ============================================================
# INSTORE SATIŞ SAYFASI
# ============================================================

@app.route("/instore-sales")
def instore_sales_page():

    return render_template(
        "instore_sales.html"
    )


# ============================================================
# MAĞAZA ARAMA
# ============================================================

@app.route(
    "/api/stores/search",
    methods=["GET"]
)
def api_search_stores():

    search = request.args.get(
        "q",
        ""
    ).strip()

    try:

        stores = search_stores(
            search
        )

        return jsonify({
            "success": True,
            "stores": stores
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error),
            "stores": []
        }), 500


# ============================================================
# PERSONEL ARAMA
# ============================================================

@app.route(
    "/api/personnel/search",
    methods=["GET"]
)
def api_search_personnel():

    store_id = request.args.get(
        "store_id",
        type=int
    )

    search = request.args.get(
        "q",
        ""
    ).strip()

    if not store_id:

        return jsonify({
            "success": False,
            "message": "store_id gereklidir.",
            "personnel": []
        }), 400

    try:

        personnel = search_personnel(
            store_id,
            search
        )

        return jsonify({
            "success": True,
            "personnel": personnel
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error),
            "personnel": []
        }), 500


# ============================================================
# MAĞAZA EXCEL AKTARIMI
# ============================================================

@app.route(
    "/api/import/stores",
    methods=["POST"]
)
def import_stores():

    try:

        file = request.files.get(
            "file"
        )

        rows = read_uploaded_file(
            file
        )

        created = 0
        updated = 0
        skipped = 0

        connection = get_connection()
        cursor = connection.cursor()

        for row in rows:

            store_code = clean_value(
                get_row_value(
                    row,
                    "Mağaza Kodu",
                    "magaza_kodu",
                    "store_code",
                    "MağazaKodu"
                )
            )

            store_name = clean_value(
                get_row_value(
                    row,
                    "Mağaza Adı",
                    "magaza_adi",
                    "store_name",
                    "MağazaAdı"
                )
            )

            city = clean_value(
                get_row_value(
                    row,
                    "İl",
                    "Il",
                    "il",
                    "city"
                )
            )

            district = clean_value(
                get_row_value(
                    row,
                    "İlçe",
                    "Ilçe",
                    "ilce",
                    "district"
                )
            )

            neighborhood = clean_value(
                get_row_value(
                    row,
                    "Mahalle",
                    "mahalle",
                    "neighborhood"
                )
            )

            if not store_code:
                skipped += 1
                continue

            if not store_name:
                skipped += 1
                continue

            cursor.execute(
                """
                SELECT id
                FROM stores
                WHERE store_code = ?
                """,
                (store_code,)
            )

            existing = cursor.fetchone()

            if existing:

                cursor.execute(
                    """
                    UPDATE stores
                    SET
                        store_name = ?,
                        city = ?,
                        district = ?,
                        neighborhood = ?
                    WHERE store_code = ?
                    """,
                    (
                        store_name,
                        city,
                        district,
                        neighborhood,
                        store_code
                    )
                )

                updated += 1

            else:

                cursor.execute(
                    """
                    INSERT INTO stores (
                        store_code,
                        store_name,
                        city,
                        district,
                        neighborhood
                    )
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        store_code,
                        store_name,
                        city,
                        district,
                        neighborhood
                    )
                )

                created += 1

        connection.commit()
        connection.close()

        return jsonify({
            "success": True,
            "message": "Mağaza aktarımı tamamlandı.",
            "created": created,
            "updated": updated,
            "skipped": skipped
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error)
        }), 400


# ============================================================
# PERSONEL EXCEL AKTARIMI
# ============================================================

@app.route(
    "/api/import/personnel",
    methods=["POST"]
)
def import_personnel():

    try:

        file = request.files.get(
            "file"
        )

        rows = read_uploaded_file(
            file
        )

        created = 0
        updated = 0
        skipped = 0

        connection = get_connection()
        cursor = connection.cursor()

        for row in rows:

            store_code = clean_value(
                get_row_value(
                    row,
                    "Mağaza Kodu",
                    "magaza_kodu",
                    "store_code"
                )
            )

            personnel_code = clean_value(
                get_row_value(
                    row,
                    "Personel Kodu",
                    "personel_kodu",
                    "personnel_code"
                )
            )

            personnel_name = clean_value(
                get_row_value(
                    row,
                    "Personel Adı",
                    "personel_adi",
                    "personnel_name"
                )
            )

            title = clean_value(
                get_row_value(
                    row,
                    "Unvan",
                    "Ünvan",
                    "title"
                )
            )

            if (
                not store_code
                or not personnel_code
                or not personnel_name
            ):

                skipped += 1
                continue

            cursor.execute(
                """
                SELECT id
                FROM stores
                WHERE store_code = ?
                """,
                (store_code,)
            )

            store = cursor.fetchone()

            if not store:

                skipped += 1
                continue

            store_id = store["id"]

            cursor.execute(
                """
                SELECT id
                FROM personnel
                WHERE personnel_code = ?
                """,
                (personnel_code,)
            )

            existing = cursor.fetchone()

            if existing:

                cursor.execute(
                    """
                    UPDATE personnel
                    SET
                        personnel_name = ?,
                        store_id = ?,
                        title = ?
                    WHERE personnel_code = ?
                    """,
                    (
                        personnel_name,
                        store_id,
                        title,
                        personnel_code
                    )
                )

                updated += 1

            else:

                cursor.execute(
                    """
                    INSERT INTO personnel (
                        personnel_code,
                        personnel_name,
                        store_id,
                        title
                    )
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        personnel_code,
                        personnel_name,
                        store_id,
                        title
                    )
                )

                created += 1

        connection.commit()
        connection.close()

        return jsonify({
            "success": True,
            "message": "Personel aktarımı tamamlandı.",
            "created": created,
            "updated": updated,
            "skipped": skipped
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error)
        }), 400


# ============================================================
# HEDEF EXCEL AKTARIMI
# ============================================================

@app.route(
    "/api/import/targets",
    methods=["POST"]
)
def import_targets():

    try:

        file = request.files.get(
            "file"
        )

        year = request.form.get(
            "year",
            type=int
        )

        week = request.form.get(
            "week",
            type=int
        )

        if not year or not week:

            return jsonify({
                "success": False,
                "message": "Yıl ve hafta seçilmelidir."
            }), 400

        rows = read_uploaded_file(
            file
        )

        count = 0
        skipped = 0

        for row in rows:

            store_code = clean_value(
                get_row_value(
                    row,
                    "Mağaza Kodu",
                    "magaza_kodu",
                    "store_code"
                )
            )

            target = get_row_value(
                row,
                "Hedef",
                "hedef",
                "target",
                "Hedef Tutar",
                "target_amount"
            )

            if (
                not store_code
                or target is None
            ):

                skipped += 1
                continue

            target_amount = normalize_number(
                target
            )

            connection = get_connection()
            cursor = connection.cursor()

            cursor.execute(
                """
                SELECT id
                FROM stores
                WHERE store_code = ?
                """,
                (store_code,)
            )

            store = cursor.fetchone()

            if not store:

                connection.close()

                skipped += 1
                continue

            save_target(
                store_id=store["id"],
                year=year,
                week=week,
                target_amount=target_amount
            )

            connection.close()

            count += 1

        return jsonify({
            "success": True,
            "message": "Hedef aktarımı tamamlandı.",
            "count": count,
            "skipped": skipped
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error)
        }), 400


# ============================================================
# PERSONEL BİREYSEL NORMAL SATIŞ EXCEL AKTARIMI
# ============================================================

@app.route(
    "/api/import/sales",
    methods=["POST"]
)
def import_sales():

    try:

        file = request.files.get(
            "file"
        )

        year = request.form.get(
            "year",
            type=int
        )

        week = request.form.get(
            "week",
            type=int
        )

        if not year or not week:

            return jsonify({
                "success": False,
                "message": "Yıl ve hafta seçilmelidir."
            }), 400

        rows = read_uploaded_file(
            file
        )

        count = 0
        skipped = 0

        connection = get_connection()
        cursor = connection.cursor()

        for row in rows:

            store_code = clean_value(
                get_row_value(
                    row,
                    "Mağaza Kodu",
                    "magaza_kodu",
                    "store_code"
                )
            )

            personnel_code = clean_value(
                get_row_value(
                    row,
                    "Personel Kodu",
                    "personel_kodu",
                    "personnel_code"
                )
            )

            amount = get_row_value(
                row,
                "Toplam Satış",
                "toplam_satis",
                "Toplam Satis",
                "Bireysel Satış",
                "bireysel_satis",
                "sales",
                "amount"
            )

            if (
                not store_code
                or not personnel_code
                or amount is None
            ):

                skipped += 1
                continue

            amount = normalize_number(
                amount
            )

            cursor.execute(
                """
                SELECT id
                FROM stores
                WHERE store_code = ?
                """,
                (store_code,)
            )

            store = cursor.fetchone()

            if not store:

                skipped += 1
                continue

            store_id = store["id"]

            cursor.execute(
                """
                SELECT id
                FROM personnel
                WHERE personnel_code = ?
                  AND store_id = ?
                """,
                (
                    personnel_code,
                    store_id
                )
            )

            personnel = cursor.fetchone()

            if not personnel:

                skipped += 1
                continue

            personnel_id = personnel["id"]

            cursor.execute(
                """
                SELECT id
                FROM sales
                WHERE year = ?
                  AND week = ?
                  AND store_id = ?
                  AND personnel_id = ?
                """,
                (
                    year,
                    week,
                    store_id,
                    personnel_id
                )
            )

            existing = cursor.fetchone()

            if existing:

                cursor.execute(
                    """
                    UPDATE sales
                    SET amount = ?
                    WHERE id = ?
                    """,
                    (
                        amount,
                        existing["id"]
                    )
                )

            else:

                cursor.execute(
                    """
                    INSERT INTO sales (
                        year,
                        week,
                        store_id,
                        personnel_id,
                        amount
                    )
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        year,
                        week,
                        store_id,
                        personnel_id,
                        amount
                    )
                )

            count += 1

        connection.commit()
        connection.close()

        return jsonify({
            "success": True,
            "message": (
                "Personel bireysel "
                "normal satış aktarımı tamamlandı."
            ),
            "count": count,
            "skipped": skipped
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error)
        }), 400


# ============================================================
# KURUMSAL SATIŞ EKLE
# ============================================================

@app.route(
    "/api/corporate-sales",
    methods=["POST"]
)
def api_create_corporate_sale():

    try:

        data = request.get_json(
            silent=True
        ) or {}

        record_id = create_corporate_sale(
            store_id=int(
                data["store_id"]
            ),
            personnel_id=int(
                data["personnel_id"]
            ),
            year=int(
                data["year"]
            ),
            week=int(
                data["week"]
            ),
            amount=float(
                data["amount"]
            ),
            description=data.get(
                "description"
            )
        )

        return jsonify({
            "success": True,
            "message": (
                "Kurumsal satış "
                "başarıyla eklendi."
            ),
            "id": record_id
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error)
        }), 400


# ============================================================
# INSTORE SATIŞ EKLE
# ============================================================

@app.route(
    "/api/instore-sales",
    methods=["POST"]
)
def api_create_instore_sale():

    try:

        data = request.get_json(
            silent=True
        ) or {}

        record_id = create_instore_sale(
            store_id=int(
                data["store_id"]
            ),
            personnel_id=int(
                data["personnel_id"]
            ),
            year=int(
                data["year"]
            ),
            week=int(
                data["week"]
            ),
            amount=float(
                data["amount"]
            ),
            description=data.get(
                "description"
            )
        )

        return jsonify({
            "success": True,
            "message": (
                "InStore satış "
                "başarıyla eklendi."
            ),
            "id": record_id
        })

    except Exception as error:

        return jsonify({
            "success": False,
            "message": str(error)
        }), 400


# ============================================================
# SAĞLIK KONTROLÜ
# ============================================================

@app.route(
    "/api/health",
    methods=["GET"]
)
def health_check():

    return jsonify({
        "success": True,
        "message": (
            "Prim hesaplama sistemi çalışıyor."
        )
    })


# ============================================================
# 404
# ============================================================

@app.errorhandler(404)
def page_not_found(error):

    return """
    <!DOCTYPE html>
    <html lang="tr">

    <head>

        <meta charset="UTF-8">

        <title>
            Sayfa Bulunamadı
        </title>

    </head>

    <body>

        <h1>
            Sayfa bulunamadı
        </h1>

        <p>
            İstenen sayfa mevcut değil.
        </p>

        <a href="/dashboard">
            Dashboard'a dön
        </a>

    </body>

    </html>
    """, 404


# ============================================================
# 500
# ============================================================

@app.errorhandler(500)
def internal_server_error(error):

    return jsonify({
        "success": False,
        "message": (
            "Sunucu tarafında bir hata oluştu."
        )
    }), 500


# ============================================================
# ÇALIŞTIR
# ============================================================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=False
    )
