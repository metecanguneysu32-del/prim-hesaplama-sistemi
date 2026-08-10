import os
import sqlite3

from database import (
    get_connection
)


# =========================================================
# GENEL AYARLAR
# =========================================================

ALLOWED_EXTENSIONS = {
    ".xlsx",
    ".xls"
}


# =========================================================
# DOSYA UZANTISI KONTROLÜ
# =========================================================

def allowed_file(filename):

    if not filename:
        return False

    extension = os.path.splitext(
        filename
    )[1].lower()

    return extension in ALLOWED_EXTENSIONS


# =========================================================
# EXCEL KÜTÜPHANESİ KONTROLÜ
# =========================================================

def get_openpyxl():

    try:

        import openpyxl

        return openpyxl

    except ImportError:

        raise RuntimeError(
            "openpyxl kütüphanesi bulunamadı."
        )


# =========================================================
# EXCEL DOSYASINI OKU
# =========================================================

def read_excel_file(
    file_path
):

    if not os.path.exists(
        file_path
    ):

        raise FileNotFoundError(
            "Excel dosyası bulunamadı."
        )


    if not allowed_file(
        file_path
    ):

        raise ValueError(
            "Sadece XLS veya XLSX dosyaları "
            "kabul edilmektedir."
        )


    openpyxl = get_openpyxl()


    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True
    )


    worksheet = workbook.active


    rows = list(
        worksheet.iter_rows(
            values_only=True
        )
    )


    workbook.close()


    if not rows:

        raise ValueError(
            "Excel dosyası boş."
        )


    headers = []

    for value in rows[0]:

        if value is None:

            headers.append("")

        else:

            headers.append(
                str(value).strip()
            )


    data_rows = rows[1:]


    return (
        headers,
        data_rows
    )


# =========================================================
# BAŞLIKLARI NORMALLEŞTİR
# =========================================================

def normalize_header(
    header
):

    if header is None:

        return ""


    value = str(
        header
    ).strip().lower()


    replacements = {

        "ı": "i",
        "İ": "i",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ş": "s",
        "Ş": "s",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c"

    }


    for old, new in replacements.items():

        value = value.replace(
            old,
            new
        )


    value = value.replace(
        " ",
        ""
    )

    value = value.replace(
        "_",
        ""
    )

    value = value.replace(
        "-",
        ""
    )


    return value


# =========================================================
# BAŞLIK BUL
# =========================================================

def find_column(
    headers,
    possible_names
):

    normalized_headers = [

        normalize_header(
            header
        )

        for header in headers

    ]


    normalized_names = [

        normalize_header(
            name
        )

        for name in possible_names

    ]


    for index, header in enumerate(
        normalized_headers
    ):

        if header in normalized_names:

            return index


    return None


# =========================================================
# DEĞERİ METNE ÇEVİR
# =========================================================

def clean_text(
    value
):

    if value is None:

        return ""


    return str(
        value
    ).strip()


# =========================================================
# SAYIYA ÇEVİR
# =========================================================

def clean_number(
    value
):

    if value is None:

        return 0.0


    if isinstance(
        value,
        (int, float)
    ):

        return float(
            value
        )


    text = str(
        value
    ).strip()


    if not text:

        return 0.0


    text = text.replace(
        "₺",
        ""
    )

    text = text.replace(
        "TL",
        ""
    )

    text = text.replace(
        "tl",
        ""
    )

    text = text.replace(
        " ",
        ""
    )


    # Türkçe Excel sayı formatı:
    #
    # 12.500,50
    #
    # -> 12500.50

    if "," in text:

        text = text.replace(
            ".",
            ""
        )

        text = text.replace(
            ",",
            "."
        )


    try:

        return float(
            text
        )

    except ValueError:

        raise ValueError(
            f"Geçersiz sayı: {value}"
        )


# =========================================================
# MAĞAZA BUL / OLUŞTUR
# =========================================================

def get_or_create_store(
    connection,
    store_code,
    store_name
):

    cursor = connection.cursor()


    cursor.execute(
        """
        SELECT id
        FROM stores
        WHERE store_code = ?
        """,
        (
            store_code,
        )
    )


    row = cursor.fetchone()


    if row:

        cursor.execute(
            """
            UPDATE stores

            SET store_name = ?

            WHERE store_code = ?
            """,
            (
                store_name,
                store_code
            )
        )


        return row["id"]


    cursor.execute(
        """
        INSERT INTO stores
        (
            store_code,
            store_name
        )

        VALUES (?, ?)
        """,
        (
            store_code,
            store_name
        )
    )


    return cursor.lastrowid


# =========================================================
# PERSONEL BUL / OLUŞTUR
# =========================================================

def get_or_create_personnel(
    connection,
    personnel_code,
    personnel_name,
    store_id,
    title=None
):

    cursor = connection.cursor()


    cursor.execute(
        """
        SELECT id
        FROM personnel
        WHERE personnel_code = ?
        """,
        (
            personnel_code,
        )
    )


    row = cursor.fetchone()


    if row:

        cursor.execute(
            """
            UPDATE personnel

            SET
                personnel_name = ?,
                store_id = ?,
                title = COALESCE(?, title)

            WHERE personnel_code = ?
            """,
            (
                personnel_name,
                store_id,
                title,
                personnel_code
            )
        )


        return row["id"]


    cursor.execute(
        """
        INSERT INTO personnel
        (
            personnel_code,
            personnel_name,
            store_id,
            title
        )

        VALUES (?, ?, ?, ?)
        """,
        (
            personnel_code,
            personnel_name,
            store_id,
            title
        )
    )


    return cursor.lastrowid


# =========================================================
# DÖNEM BUL / OLUŞTUR
# =========================================================

def get_or_create_period(
    connection,
    year,
    week
):

    cursor = connection.cursor()


    cursor.execute(
        """
        SELECT id
        FROM periods

        WHERE
            year = ?
            AND week = ?
        """,
        (
            year,
            week
        )
    )


    row = cursor.fetchone()


    if row:

        return row["id"]


    name = (
        f"{year} - {week}. Hafta"
    )


    cursor.execute(
        """
        INSERT INTO periods
        (
            year,
            week,
            name
        )

        VALUES (?, ?, ?)
        """,
        (
            year,
            week,
            name
        )
    )


    return cursor.lastrowid


# =========================================================
# MAĞAZA EXCEL AKTARIMI
# =========================================================

def import_stores(
    file_path
):

    headers, rows = read_excel_file(
        file_path
    )


    store_code_index = find_column(
        headers,
        [
            "Mağaza Kodu",
            "Magaza Kodu",
            "MağazaKod",
            "MagazaKod",
            "Store Code",
            "StoreCode"
        ]
    )


    store_name_index = find_column(
        headers,
        [
            "Mağaza Adı",
            "Magaza Adi",
            "Mağaza",
            "Magaza",
            "Store Name",
            "StoreName"
        ]
    )


    if store_code_index is None:

        raise ValueError(
            "Excel dosyasında "
            "'Mağaza Kodu' sütunu bulunamadı."
        )


    if store_name_index is None:

        raise ValueError(
            "Excel dosyasında "
            "'Mağaza Adı' sütunu bulunamadı."
        )


    connection = get_connection()


    success_rows = 0

    error_rows = 0

    errors = []


    try:

        for row_number, row in enumerate(
            rows,
            start=2
        ):

            try:

                store_code = clean_text(
                    row[store_code_index]
                )


                store_name = clean_text(
                    row[store_name_index]
                )


                if not store_code:

                    raise ValueError(
                        "Mağaza kodu boş."
                    )


                if not store_name:

                    raise ValueError(
                        "Mağaza adı boş."
                    )


                get_or_create_store(
                    connection,
                    store_code,
                    store_name
                )


                success_rows += 1


            except Exception as error:

                error_rows += 1

                errors.append({

                    "row": row_number,

                    "error": str(error)

                })


        connection.commit()


    except Exception:

        connection.rollback()

        raise


    finally:

        connection.close()


    return {

        "total_rows": len(rows),

        "success_rows": success_rows,

        "error_rows": error_rows,

        "errors": errors

    }


# =========================================================
# PERSONEL EXCEL AKTARIMI
# =========================================================

def import_personnel(
    file_path
):

    headers, rows = read_excel_file(
        file_path
    )


    store_code_index = find_column(
        headers,
        [
            "Mağaza Kodu",
            "Magaza Kodu",
            "Store Code",
            "StoreCode"
        ]
    )


    store_name_index = find_column(
        headers,
        [
            "Mağaza Adı",
            "Magaza Adi",
            "Store Name",
            "StoreName"
        ]
    )


    personnel_code_index = find_column(
        headers,
        [
            "Personel Kodu",
            "Personel Kod",
            "Personel No",
            "Personel Numarası",
            "Personnel Code",
            "PersonnelCode"
        ]
    )


    personnel_name_index = find_column(
        headers,
        [
            "Personel Adı",
            "Personel Ad Soyad",
            "Personel",
            "Personnel Name",
            "PersonnelName"
        ]
    )


    title_index = find_column(
        headers,
        [
            "Unvan",
            "Ünvan",
            "Görev",
            "Pozisyon",
            "Title"
        ]
    )


    if store_code_index is None:

        raise ValueError(
            "'Mağaza Kodu' sütunu bulunamadı."
        )


    if store_name_index is None:

        raise ValueError(
            "'Mağaza Adı' sütunu bulunamadı."
        )


    if personnel_code_index is None:

        raise ValueError(
            "'Personel Kodu' sütunu bulunamadı."
        )


    if personnel_name_index is None:

        raise ValueError(
            "'Personel Adı' sütunu bulunamadı."
        )


    connection = get_connection()


    success_rows = 0

    error_rows = 0

    errors = []


    try:

        for row_number, row in enumerate(
            rows,
            start=2
        ):

            try:

                store_code = clean_text(
                    row[store_code_index]
                )


                store_name = clean_text(
                    row[store_name_index]
                )


                personnel_code = clean_text(
                    row[personnel_code_index]
                )


                personnel_name = clean_text(
                    row[personnel_name_index]
                )


                title = None


                if title_index is not None:

                    title = clean_text(
                        row[title_index]
                    )


                if not store_code:

                    raise ValueError(
                        "Mağaza kodu boş."
                    )


                if not store_name:

                    raise ValueError(
                        "Mağaza adı boş."
                    )


                if not personnel_code:

                    raise ValueError(
                        "Personel kodu boş."
                    )


                if not personnel_name:

                    raise ValueError(
                        "Personel adı boş."
                    )


                store_id = get_or_create_store(
                    connection,
                    store_code,
                    store_name
                )


                get_or_create_personnel(
                    connection,
                    personnel_code,
                    personnel_name,
                    store_id,
                    title
                )


                success_rows += 1


            except Exception as error:

                error_rows += 1

                errors.append({

                    "row": row_number,

                    "error": str(error)

                })


        connection.commit()


    except Exception:

        connection.rollback()

        raise


    finally:

        connection.close()


    return {

        "total_rows": len(rows),

        "success_rows": success_rows,

        "error_rows": error_rows,

        "errors": errors

    }


# =========================================================
# HEDEF EXCEL AKTARIMI
# =========================================================

def import_targets(
    file_path,
    year,
    week
):

    headers, rows = read_excel_file(
        file_path
    )


    store_code_index = find_column(
        headers,
        [
            "Mağaza Kodu",
            "Magaza Kodu",
            "Store Code",
            "StoreCode"
        ]
    )


    target_index = find_column(
        headers,
        [
            "Hedef",
            "Hedef Tutar",
            "Haftalık Hedef",
            "Target",
            "Target Amount"
        ]
    )


    if store_code_index is None:

        raise ValueError(
            "'Mağaza Kodu' sütunu bulunamadı."
        )


    if target_index is None:

        raise ValueError(
            "'Hedef' sütunu bulunamadı."
        )


    connection = get_connection()


    success_rows = 0

    error_rows = 0

    errors = []


    try:

        period_id = get_or_create_period(
            connection,
            year,
            week
        )


        for row_number, row in enumerate(
            rows,
            start=2
        ):

            try:

                store_code = clean_text(
                    row[store_code_index]
                )


                target_amount = clean_number(
                    row[target_index]
                )


                if not store_code:

                    raise ValueError(
                        "Mağaza kodu boş."
                    )


                if target_amount < 0:

                    raise ValueError(
                        "Hedef negatif olamaz."
                    )


                cursor = connection.cursor()


                cursor.execute(
                    """
                    SELECT id
                    FROM stores
                    WHERE store_code = ?
                    """,
                    (
                        store_code,
                    )
                )


                store = cursor.fetchone()


                if not store:

                    raise ValueError(
                        f"Mağaza bulunamadı: "
                        f"{store_code}"
                    )


                store_id = store["id"]


                cursor.execute(
                    """
                    INSERT INTO store_targets
                    (
                        store_id,
                        period_id,
                        target_amount
                    )

                    VALUES (?, ?, ?)

                    ON CONFLICT(
                        store_id,
                        period_id
                    )

                    DO UPDATE SET

                        target_amount =
                            excluded.target_amount
                    """,
                    (
                        store_id,
                        period_id,
                        target_amount
                    )
                )


                success_rows += 1


            except Exception as error:

                error_rows += 1

                errors.append({

                    "row": row_number,

                    "error": str(error)

                })


        connection.commit()


    except Exception:

        connection.rollback()

        raise


    finally:

        connection.close()


    return {

        "total_rows": len(rows),

        "success_rows": success_rows,

        "error_rows": error_rows,

        "errors": errors

    }


# =========================================================
# SATIŞ EXCEL AKTARIMI
# =========================================================
#
# Beklenen temel Excel:
#
# Mağaza Kodu
# Mağaza Adı
# Personel Kodu
# Personel Adı
# Toplam Satış
#
# Bir personel bir mağazada çalışır.
# Normal satış burada tek satır olarak tutulur.
# Kurumsal ve InStore daha sonra ayrı ekranlardan
# sisteme eklenecektir.
# =========================================================

def import_sales(
    file_path,
    year,
    week
):

    headers, rows = read_excel_file(
        file_path
    )


    store_code_index = find_column(
        headers,
        [
            "Mağaza Kodu",
            "Magaza Kodu",
            "Store Code",
            "StoreCode"
        ]
    )


    store_name_index = find_column(
        headers,
        [
            "Mağaza Adı",
            "Magaza Adi",
            "Store Name",
            "StoreName"
        ]
    )


    personnel_code_index = find_column(
        headers,
        [
            "Personel Kodu",
            "Personel Kod",
            "Personel No",
            "Personnel Code",
            "PersonnelCode"
        ]
    )


    personnel_name_index = find_column(
        headers,
        [
            "Personel Adı",
            "Personel Ad Soyad",
            "Personel",
            "Personnel Name",
            "PersonnelName"
        ]
    )


    sales_index = find_column(
        headers,
        [
            "Toplam Satış",
            "Toplam Satis",
            "Satış",
            "Satis",
            "Satış Tutarı",
            "Satis Tutari",
            "Total Sales",
            "Sales"
        ]
    )


    if store_code_index is None:

        raise ValueError(
            "'Mağaza Kodu' sütunu bulunamadı."
        )


    if store_name_index is None:

        raise ValueError(
            "'Mağaza Adı' sütunu bulunamadı."
        )


    if personnel_code_index is None:

        raise ValueError(
            "'Personel Kodu' sütunu bulunamadı."
        )


    if personnel_name_index is None:

        raise ValueError(
            "'Personel Adı' sütunu bulunamadı."
        )


    if sales_index is None:

        raise ValueError(
            "'Toplam Satış' sütunu bulunamadı."
        )


    connection = get_connection()


    success_rows = 0

    error_rows = 0

    errors = []


    try:

        period_id = get_or_create_period(
            connection,
            year,
            week
        )


        for row_number, row in enumerate(
            rows,
            start=2
        ):

            try:

                store_code = clean_text(
                    row[store_code_index]
                )


                store_name = clean_text(
                    row[store_name_index]
                )


                personnel_code = clean_text(
                    row[personnel_code_index]
                )


                personnel_name = clean_text(
                    row[personnel_name_index]
                )


                sales_amount = clean_number(
                    row[sales_index]
                )


                if not store_code:

                    raise ValueError(
                        "Mağaza kodu boş."
                    )


                if not store_name:

                    raise ValueError(
                        "Mağaza adı boş."
                    )


                if not personnel_code:

                    raise ValueError(
                        "Personel kodu boş."
                    )


                if not personnel_name:

                    raise ValueError(
                        "Personel adı boş."
                    )


                if sales_amount < 0:

                    raise ValueError(
                        "Satış tutarı negatif olamaz."
                    )


                store_id = get_or_create_store(
                    connection,
                    store_code,
                    store_name
                )


                personnel_id = (
                    get_or_create_personnel(
                        connection,
                        personnel_code,
                        personnel_name,
                        store_id
                    )
                )


                cursor = connection.cursor()


                cursor.execute(
                    """
                    INSERT INTO personnel_sales
                    (
                        personnel_id,
                        store_id,
                        period_id,
                        sales_amount
                    )

                    VALUES (?, ?, ?, ?)

                    ON CONFLICT(
                        personnel_id,
                        period_id
                    )

                    DO UPDATE SET

                        store_id =
                            excluded.store_id,

                        sales_amount =
                            excluded.sales_amount
                    """,
                    (
                        personnel_id,
                        store_id,
                        period_id,
                        sales_amount
                    )
                )


                success_rows += 1


            except Exception as error:

                error_rows += 1

                errors.append({

                    "row": row_number,

                    "error": str(error)

                })


        connection.commit()


    except Exception:

        connection.rollback()

        raise


    finally:

        connection.close()


    return {

        "total_rows": len(rows),

        "success_rows": success_rows,

        "error_rows": error_rows,

        "errors": errors

    }


# =========================================================
# IMPORT LOG KAYDI
# =========================================================

def save_import_log(
    file_type,
    file_name,
    total_rows,
    success_rows,
    error_rows
):

    connection = get_connection()


    try:

        cursor = connection.cursor()


        cursor.execute(
            """
            INSERT INTO import_logs
            (
                file_type,
                file_name,
                total_rows,
                success_rows,
                error_rows
            )

            VALUES (?, ?, ?, ?, ?)
            """,
            (
                file_type,
                file_name,
                total_rows,
                success_rows,
                error_rows
            )
        )


        connection.commit()


    finally:

        connection.close()


# =========================================================
# HATA EXCELİ OLUŞTUR
# =========================================================

def create_error_excel(
    errors,
    output_path
):

    openpyxl = get_openpyxl()


    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.title = "Hatalar"


    worksheet.append(
        [
            "Excel Satırı",
            "Hata"
        ]
    )


    for error in errors:

        worksheet.append(
            [
                error.get(
                    "row",
                    ""
                ),

                error.get(
                    "error",
                    ""
                )
            ]
        )


    worksheet.column_dimensions[
        "A"
    ].width = 15


    worksheet.column_dimensions[
        "B"
    ].width = 70


    workbook.save(
        output_path
    )


    workbook.close()


    return output_path